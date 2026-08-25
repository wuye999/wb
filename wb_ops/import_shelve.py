# -*- coding: utf-8 -*-
"""
wb_ops 他人映射表导入上架（import-shelve）

把他人（同项目格式）映射表中有、我方 5 店全无的商品，上架到我的店铺：
- 差集判断：按 WB原始nmId（vc 末段数字）匹配，两代 vc 格式（随机4位/前缀码）通吃。
- 新 vc 生成：我方前缀优先——中文名精确命中我商品价格表前缀码 → BCS-{我的前缀}-{nmId}；
  未命中 → 沿用他人 vc 原样。
- 价格：他人表「双倍售价」列，店铺价 = floor(双倍售价)（我方标准定价规则）。
- 商品数据：他人表（俄文标题/主图/尺寸/毛重）+ WB detail（BCS 代理）+ card.json（CDN）兜底。
- 防重复：我方快照 nm 集合判断 + 本地记录（nm 键）+ 执行时 vendorCodeMulti 实时查重。
  （注意：新 vc 与他人 vc 不同时 BCS 服务端幂等失效，本地记录是必要防线。）

用法：wb.py import-shelve <他人映射表.xlsx> [--cn|--shops|--limit] [--apply] [--no-verify] [--interval S]
"""
import copy
import csv
import json
import math
import os
import time
from datetime import datetime

import openpyxl

from . import bcs
from . import config
from . import products
from . import replicate

# 本地提交记录（与 replicate 共用文件；键 = nmId，比 vc 稳定——同一商品可能两种 vc 格式）
RECORDS_JSON = replicate.RECORDS_JSON

# 他人表「映射总表」Sheet 用到的列名（按表头定位，缺列容错）
COL_CN = "产品中文名"
COL_VC = "vendorCode"
COL_DP = "双倍售价"
COL_TITLE = "俄文标题"
COL_IMG = "主图链接"
COL_L = "尺寸长(cm)"
COL_W = "尺寸宽(cm)"
COL_H = "尺寸高(cm)"
COL_WT = "毛重(kg)"
COL_NM = "WB商品码"  # 他人店铺可靠 nmId（新格式）；旧格式表无此列

# 上架库存按中文名决定：复用 replicate.stock_for（--cn-stock 指定，未指定默认 999）


def nm_of(vc):
    """vc → WB原始nmId（末段纯数字才有效，否则 None）"""
    tail = str(vc or "").rsplit("-", 1)[-1]
    return tail if tail.isdigit() else None


# ---------------- 包装数据：商品价格表兜底 ----------------
_pkg_cache = None


def boss_pkg_map():
    """商品价格表「尺寸」列 → {中文名: (L, W, H, weight)}（格式 `长*宽*高/毛重`，容错空格）。
    兜底来源：他人表缺毛重/尺寸时，按中文名取我方商品价格表同名的包装数据。"""
    global _pkg_cache
    if _pkg_cache is not None:
        return _pkg_cache
    import re as _re
    from . import mapping
    out = {}
    try:
        wb = openpyxl.load_workbook(config.BOSS_XLSX, data_only=True)
        ws = wb["Sheet1"]
        for r in ws.iter_rows(min_row=2, values_only=True):
            cn = str(r[2] or "").strip()
            dim = str(r[4] or "").strip() if len(r) > 4 else ""
            if not cn or not dim:
                continue
            m = _re.match(r"^([\d.]+)\s*\*\s*([\d.]+)\s*\*\s*([\d.]+)\s*/\s*([\d.]+)$", dim)
            if not m:
                continue
            out[cn] = (float(m.group(1)), float(m.group(2)),
                       float(m.group(3)), float(m.group(4)))
        wb.close()
    except Exception:
        pass
    _pkg_cache = out
    return out


# ---------------- 他人表解析 ----------------
def load_foreign(xlsx_path):
    """解析他人映射表「映射总表」Sheet → (items, bad_vcs)
    items: [{cn, vc, k, nm, dp, title_ru, img, L, W, H, weight}]，同 k 去重（优先双倍售价非空行）
      k：= vc 末段(WB原始nmId)，跨账号稳定匹配键（差集/去重/新 vc 尾段）
      nm：= 他人表「WB商品码」列(他人店铺 nmId)，可靠拉取标识；旧格式表无此列 → 为 None（上架时跳过）
    bad_vcs: 末段非数字的 vc（格式异常，跳过）"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "映射总表" not in wb.sheetnames:
        wb.close()
        raise RuntimeError(f"{xlsx_path} 缺少「映射总表」Sheet（需同项目生成的映射表格式）")
    ws = wb["映射总表"]
    headers = [str(c.value or "") for c in ws[1]]

    def idx(name):
        return headers.index(name) if name in headers else None

    i_cn, i_vc, i_dp = idx(COL_CN), idx(COL_VC), idx(COL_DP)
    i_title, i_img = idx(COL_TITLE), idx(COL_IMG)
    i_l, i_w, i_h, i_wt = idx(COL_L), idx(COL_W), idx(COL_H), idx(COL_WT)
    i_nm = idx(COL_NM)
    if i_vc is None or i_cn is None:
        wb.close()
        raise RuntimeError(f"表头缺少 {COL_VC}/{COL_CN} 列，无法解析")
    if i_nm is None:
        print(f"[提醒] 他人表缺少「{COL_NM}」列（旧格式），无法取可靠 nmId → 将全部跳过上架，请他人用新格式导出")

    by_k, bad_vcs = {}, []
    for r in ws.iter_rows(min_row=2, values_only=True):
        vc = str(r[i_vc] or "").strip()
        cn = str(r[i_cn] or "").strip() if i_cn is not None else ""
        if not vc:
            continue
        k = nm_of(vc)
        if k is None:
            bad_vcs.append(vc)
            continue
        raw_nm = r[i_nm] if i_nm is not None else None
        nm = str(raw_nm).strip() if raw_nm not in (None, "") else None
        item = {
            "cn": cn, "vc": vc, "k": k, "nm": nm,
            "dp": r[i_dp] if i_dp is not None else None,
            "title_ru": str(r[i_title] or "").strip() if i_title is not None else "",
            "img": str(r[i_img] or "").strip() if i_img is not None else "",
            "L": r[i_l] if i_l is not None else None,
            "W": r[i_w] if i_w is not None else None,
            "H": r[i_h] if i_h is not None else None,
            "weight": r[i_wt] if i_wt is not None else None,
        }
        old = by_k.get(k)
        if old is None or (old["dp"] is None and item["dp"] is not None):
            by_k[k] = item  # 同 k 多行：优先保留双倍售价非空者
    wb.close()
    return list(by_k.values()), bad_vcs


# ---------------- 差集与新 vc ----------------
def diff_foreign(foreign):
    """他人表 - 我方快照 → 候选清单 [(item, new_vc, prefix_from)]。
    prefix_from: '我方'/'他人'；new_vc 按我方前缀优先规则生成。"""
    from . import mapping  # 延迟 import
    shops_data, _ = products.load_all_shops()
    vc_shops, _, all_shop_ids = replicate.build_coverage(shops_data)
    my_nms = {nm_of(vc) for vc in vc_shops} - {None}

    # 我方前缀：{中文名: 前缀}（load_prefix_map 为 {前缀: {sku,cn,dp}}，反查取第一个）
    pmap = mapping.load_prefix_map()
    cn2prefix = {}
    for prefix, info in pmap.items():
        cn2prefix.setdefault(info.get("cn") or "", prefix)

    records = replicate._load_records()
    plans, have_cnt, rec_skip = [], 0, 0
    for item in sorted(foreign, key=lambda x: x["k"]):
        if item["k"] in my_nms:
            have_cnt += 1
            continue
        if item["k"] in records:  # 本地记录（键=k，跨账号稳定）已提交过
            rec_skip += 1
            continue
        prefix = cn2prefix.get(item["cn"])
        if prefix:
            new_vc, prefix_from = f"BCS-{prefix}-{item['k']}", "我方"
        else:
            new_vc, prefix_from = item["vc"], "他人"
        plans.append((item, new_vc, prefix_from))
    return plans, all_shop_ids, have_cnt, rec_skip


# ---------------- 上架请求体 ----------------
def build_import_payload(item, new_vc, sid, warehouse_id, detail, card_info, overrides=None):
    """构造 /system/wbCollection/wb/new 请求体（单目标店）。返回 (payload, err)。
    数据源：他人表列 → WB detail / card.json 兜底；价格 = floor(双倍售价)；sizes 只取第 1 条。
    overrides：中文名→库存覆盖（--cn-stock）。"""
    nm_id = int(item["nm"])
    # 价格：floor(双倍售价)
    if item["dp"] is None or float(item["dp"]) <= 0:
        return None, "双倍售价缺失或无效"
    price_str = f"{math.floor(float(item['dp']))}.00"

    # wbDetail：detail 深拷贝 + sizes 只留第 1 条（他人无规格信息，保守防误上兄弟颜色）
    wb_detail = copy.deepcopy(detail)
    sizes = wb_detail.get("sizes") or []
    if not sizes:
        return None, "WB 详情缺少规格数据（sizes 为空）"
    s0 = sizes[0]
    orig = s0.get("origName")
    if orig is None:
        orig = s0.get("techSize") if s0.get("techSize") else 0
    try:
        orig = int(float(orig))
    except (TypeError, ValueError):
        orig = 0
    wb_detail["sizes"] = [{"origName": orig,
                           "name": s0.get("name") if s0.get("name") is not None else (s0.get("wbSize") or ""),
                           "price": price_str,
                           "vendorCode": new_vc}]

    # 图片：card.json photo_count 生成全集；主图他人表优先
    images = ";".join(replicate.generate_image_urls(nm_id, ((card_info or {}).get("media") or {}).get("photo_count")))
    main_image = item["img"] or images.split(";")[0]

    # 包装尺寸/重量：他人表列 → card.json 包装组兜底 → 商品价格表同名中文名兜底 → 仍缺失败
    pkg = replicate.parse_package_info(card_info)
    length = item["L"] or pkg["length"]
    width = item["W"] or pkg["width"]
    height = item["H"] or pkg["height"]
    weight = item["weight"] or pkg["weight"]
    if not (length and width and height and weight):
        bp = boss_pkg_map().get(item["cn"] or "")
        if bp:
            bl, bw, bh, bwgt = bp
            length = length or bl
            width = width or bw
            height = height or bh
            weight = weight or bwgt
    if not length or not width or not height or not weight:
        return None, f"包装数据缺失（L={length} W={width} H={height} 重={weight}）"

    # 标题/类目：他人表俄文标题 → detail.name；类目仅 detail 有
    title = item["title_ru"] or detail.get("name") or ""
    if not title:
        return None, "商品标题缺失（他人表俄文标题与 WB 详情 name 均空）"
    subject_id = detail.get("subjectId")
    if not subject_id:
        return None, "类目 ID 缺失（WB 详情无 subjectId）"
    parent_id = detail.get("subjectParentId") or ((card_info or {}).get("data") or {}).get("subject_root_id") or 0
    parent_name = (card_info or {}).get("subj_root_name") or ""

    shop_arr = [{"id": sid, "warehouseId": warehouse_id,
                 "warehouseQuantity": replicate.stock_for(item["cn"], overrides)}]
    payload = {
        "shop": json.dumps(shop_arr, ensure_ascii=False, separators=(",", ":")),
        "offerDIY": "",
        "brandStatus": False,
        "oModel": True,
        "status": "1",
        "collectionType": "0",
        "shopDatas": [{
            "nmId": nm_id,
            "wbDetail": json.dumps(wb_detail, ensure_ascii=False, separators=(",", ":")),
            "name": title,
            "subjectId": subject_id,
            "parentId": parent_id,
            "wbCardInfo": json.dumps(card_info, ensure_ascii=False, separators=(",", ":")),
            "images": images,
            "video": "",
            "mainImage": main_image,
            "packagLength": math.ceil(float(length)),
            "packagWidth": math.ceil(float(width)),
            "packagHeight": math.ceil(float(height)),
            "weightBrutto": str(weight),
            "vendorCode": new_vc,
            "brand": "",
            "hsCode": None,
            "subjectName": detail.get("subjectName") or "",
            "parentName": parent_name,
            "sourceSku": str(nm_id),
        }],
        "mode": 1,
        "mergeCards": 1,
        "carryBrand": 1,
        "customBrand": None,
        "titleSuffix": None,
        "aiRewrite": False,
        "aiRetouch": False,
        "aiRetouchTemplateId": None,
        "imgUploadMode": 0,
    }
    return payload, None


# ---------------- 主流程 ----------------
def run(args):
    overrides = replicate.parse_cn_stock(getattr(args, "cn_stock", "") or "")
    # 前置自动同步：直接刷新全部店铺快照，保证差集判断基于最新数据
    replicate.ensure_snapshots(args)
    foreign, bad_vcs = load_foreign(args.xlsx)
    plans, all_shop_ids, have_cnt, rec_skip = diff_foreign(foreign)

    # --cn 过滤（他人表中文名包含匹配，逗号分隔多个）/ --shops 限定目标店 / --limit
    if getattr(args, "cn", ""):
        kws = [k.strip() for k in args.cn.split(",") if k.strip()]
        plans = [p for p in plans if any(kw in (p[0]["cn"] or "") for kw in kws)]
    allow_shops = ([int(x) for x in args.shops.split(",") if x.strip()]
                   if getattr(args, "shops", "") else None)
    target_shops = allow_shops or all_shop_ids
    if getattr(args, "limit", 0):
        plans = plans[:args.limit]

    print(f"他人表：{len(foreign)} 个唯一商品（格式异常跳过 {len(bad_vcs)} 个 vc）")
    if bad_vcs:
        print("  [格式异常] " + ", ".join(bad_vcs[:10]) + ("..." if len(bad_vcs) > 10 else ""))
    no_nm = [p for p in plans if not p[0].get("nm")]
    print(f"我方已有 {have_cnt} | 本地记录已提交 {rec_skip} | 待上架候选 {len(plans)} 个"
          f"（其中 {len(no_nm)} 个无WB商品码将跳过）")
    print(f"目标店铺: {target_shops}")
    if no_nm:
        print("[无WB商品码跳过] "
              + ", ".join(f"{p[0]['k']}({p[0]['cn'] or '-'})" for p in no_nm[:10])
              + ("..." if len(no_nm) > 10 else ""))

    # dry-run 清单
    print("\n" + "=" * 126)
    print(f"{'vc末段nmId':<13}{'中文名':<16}{'双倍售价→店铺价':<16}{'库存':<6}{'新vendorCode(前缀来源)':<32}{'他人vc':<26}目标店")
    print("-" * 126)
    for item, new_vc, prefix_from in plans:
        dp = item["dp"]
        shop_price = math.floor(float(dp)) if dp is not None and float(dp) > 0 else "?"
        stk = replicate.stock_for(item["cn"], overrides)
        flag = "  [跳过-无WB商品码]" if not item.get("nm") else ""
        print(f"{item['k']:<13}{(item['cn'] or '-'):<16}{f'{dp}→{shop_price}' if dp else '?':<16}"
              f"{stk:<6}{new_vc + '(' + prefix_from + ')':<32}{item['vc']:<26}{','.join(map(str, target_shops))}{flag}")
    print("=" * 126)

    if not plans:
        print("无可执行任务")
        return 0
    if not args.apply:
        print(f"\n[dry-run] 共 {len(plans)} 个商品待上架。加 --apply 执行。")
        return 0

    # ---- 执行 ----
    print(f"\n开始执行：{len(plans)} 个商品 ...")
    shops_data, _ = products.load_all_shops()
    warehouses = {}
    for sid in target_shops:
        wh = replicate.main_warehouse(sid, shops_data)
        if wh is None:
            print(f"  [警告] 店 {sid} 无法解析主仓库，该店目标将跳过")
        else:
            warehouses[sid] = wh
    print(f"各店主仓库: {warehouses}")
    if not warehouses:
        print("[错误] 无可用目标仓库")
        return 1

    records = replicate._load_records()
    csv_path = os.path.join(config.LOG_DIR, f"导入上架_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv")
    os.makedirs(config.LOG_DIR, exist_ok=True)
    csv_file = open(csv_path, "w", newline="", encoding="utf-8-sig")
    writer = csv.writer(csv_file)
    writer.writerow(["时间", "nmId", "中文名", "新vendorCode", "前缀来源", "他人vc", "目标店", "店铺价", "库存", "结果", "原因"])

    ok = skip = fail = 0
    detail_fail_streak = 0
    aborted = False
    t0 = time.time()
    for i, (item, new_vc, prefix_from) in enumerate(plans, 1):
        tag = f"[{i}/{len(plans)}]"
        k, cn = item["k"], item["cn"] or "-"
        nm = item.get("nm")  # 可靠 WB商品码（他人店铺 nmId）
        dp = item["dp"]
        shop_price = math.floor(float(dp)) if dp is not None and float(dp) > 0 else "?"
        stk = replicate.stock_for(item["cn"], overrides)
        now = datetime.now().strftime("%H:%M:%S")
        # 0) 可靠 WB商品码缺失则跳过，避免抓取过期 WB原始nmId 误上架
        if not nm:
            print(f"  {tag} [跳过] nm={k} 他人表无WB商品码，避免误上架")
            writer.writerow([now, k, cn, new_vc, prefix_from, item["vc"],
                             ",".join(map(str, target_shops)), shop_price, stk, "跳过", "他人表无WB商品码（避免误上架）"])
            skip += 1
            continue
        if aborted:
            writer.writerow([now, k, cn, new_vc, prefix_from, item["vc"],
                             ",".join(map(str, target_shops)), shop_price, stk, "中止", "连续 detail 失败触发中止"])
            skip += 1
            continue

        # 1) 查重：本地记录（键=k，跨账号稳定）+ 实时 API（按新 vc）→ 得本次可提交店
        valid = []
        for sid in target_shops:
            if sid not in warehouses:
                continue
            if str(sid) in (records.get(k) or {}):
                print(f"  {tag} [跳过] 店{sid} 本地记录已提交 nm={k}")
                continue
            if replicate.vc_exists_in_shop(new_vc, sid):
                print(f"  {tag} [跳过] 店{sid} 已存在 {new_vc}")
                continue
            valid.append(sid)
        if not valid:
            print(f"  {tag} [跳过] nm={k} 目标店均已存在或无仓库")
            writer.writerow([now, k, cn, new_vc, prefix_from, item["vc"],
                             ",".join(map(str, target_shops)), shop_price, stk, "跳过", "已存在或无仓库"])
            skip += 1
            continue

        # 2) WB detail：按 --detail-source 选通道
        #    synthetic=他人表行 + card.json 拼装（不依赖 WB detail）；auto=BCS 代理→拼装兜底
        ds = getattr(args, "detail_source", "auto")
        card_info = None
        if ds == "synthetic":
            card_info = replicate.fetch_card_json(nm)
            time.sleep(replicate.CARD_INTERVAL)
            detail = replicate.build_synthetic_detail(item, card_info, nm) if card_info else None
        else:
            detail = replicate.fetch_wb_detail(nm)
            time.sleep(replicate.DETAIL_INTERVAL)
            if detail is None and ds == "auto":  # auto 兜底：BCS 拼装
                card_info = replicate.fetch_card_json(nm)
                time.sleep(replicate.CARD_INTERVAL)
                if card_info is not None:
                    detail = replicate.build_synthetic_detail(item, card_info, nm)
        if detail is None:
            # ★ synthetic 模式失败=该商品 card.json 缺失（个体坏数据），非反爬限流，不累计中止计数
            if ds != "synthetic":
                detail_fail_streak += 1
                if detail_fail_streak >= replicate.DETAIL_FAIL_ABORT:
                    print("  [中止] 连续多个商品 detail 失败，疑似反爬限流，停止后续执行（已成功不回滚）")
                    aborted = True
            print(f"  {tag} [失败] nm={nm} WB detail 获取失败")
            writer.writerow([now, nm, cn, new_vc, prefix_from, item["vc"],
                             ",".join(map(str, valid)), shop_price, stk, "失败", "WB detail 获取失败"])
            fail += 1
            continue
        detail_fail_streak = 0

        # 3) card.json（CDN；synthetic/auto 兜底已取则跳过）
        if card_info is None:
            card_info = replicate.fetch_card_json(nm)
            time.sleep(replicate.CARD_INTERVAL)
        if card_info is None:
            print(f"  {tag} [失败] nm={nm} card.json 获取失败")
            writer.writerow([now, nm, cn, new_vc, prefix_from, item["vc"],
                             ",".join(map(str, valid)), shop_price, stk, "失败", "card.json 获取失败"])
            fail += 1
            continue

        # 4) 构造 + 逐店提交（BCS 多店 bug：必须每店单独请求）
        for sid in valid:
            payload, err = build_import_payload(item, new_vc, sid, warehouses[sid], detail, card_info, overrides)
            if payload is None:
                fail += 1
                print(f"  {tag} [失败] nm={nm} 店{sid}: {err}")
                writer.writerow([now, nm, cn, new_vc, prefix_from, item["vc"], str(sid), shop_price, stk, "失败", err])
                continue
            try:
                resp = bcs.http_post_json(f"{bcs.base_url()}/system/wbCollection/wb/new", payload)
                if resp.get("code") == 200:
                    ok += 1
                    stk_tag = "（0库存）" if stk == 0 else ""
                    print(f"  {tag} [成功] {new_vc} → 店{sid}（¥{shop_price}，{prefix_from}前缀{stk_tag}）")
                    writer.writerow([now, nm, cn, new_vc, prefix_from, item["vc"], str(sid), shop_price, stk, "成功", ""])
                    records.setdefault(k, {})[str(sid)] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    replicate._save_records(records)
                else:
                    fail += 1
                    msg = resp.get("msg") or f"code={resp.get('code')}"
                    print(f"  {tag} [失败] nm={nm} 店{sid}: {msg}")
                    writer.writerow([now, nm, cn, new_vc, prefix_from, item["vc"], str(sid), shop_price, stk, "失败", msg])
            except Exception as e:
                fail += 1
                print(f"  {tag} [失败] nm={nm} 店{sid}: {e}")
                writer.writerow([now, nm, cn, new_vc, prefix_from, item["vc"], str(sid), shop_price, stk, "失败", str(e)])
            time.sleep(args.interval)
        if i < len(plans):
            time.sleep(args.interval)

    csv_file.close()
    print(f"\n[汇总] 计划 {len(plans)} | 成功 {ok} | 跳过 {skip} | 失败 {fail}（{time.time() - t0:.0f}s）")
    print(f"明细：{csv_path}")

    # ---- 写后验证：仅加 --sync 时 fetch 同步复核（否则只打印提示，不自动做）----
    if ok > 0 and getattr(args, "sync", False) and not args.no_verify:
        print("\n[写后验证] 触发全店同步 + 拉取（~1.5 分钟）...")
        try:
            products.fetch_all()
            shops_data2, _ = products.load_all_shops()
            vc_shops2, _, _ = replicate.build_coverage(shops_data2)
            per_shop = {sid: len([1 for s in vc_shops2.values() if sid in s]) for sid in all_shop_ids}
            print(f"[验证] 各店在架 vc 数：{per_shop}")
        except Exception as e:
            print(f"[验证] 失败：{e}（可稍后手动 wb.py fetch 复核）")
        # 写后验证已 fetch 最新快照 → 顺带增量合并映射表（上架后自动补录/同步覆盖）
        from . import mapping_sync
        mapping_sync.post_write_merge(fetch=False)
    elif ok > 0:
        from . import mapping_sync
        mapping_sync.print_write_hint()
    return 0
