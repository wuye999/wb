# -*- coding: utf-8 -*-
"""
wb_ops 多店铺同步 / 待审核 / 增量合并（原 sync_shops.py）

review：找出其他店铺在架但映射表中没有的 vc（前缀命中→自动补录，其余→待审核）。
merge：增量合并（映射表 xlsx = 唯一状态源，继承旧归属 + 追加审核 + 消失即移除）。
"""
import json
import os
import re

from . import config
from . import mapping
from . import products
from . import workbench
VC_PREFIX_RE = re.compile(config.VC_PREFIX_RE)


def load_known_vcs():
    """已知 vendorCode = 映射表「映射总表」已归属 vc + 「已排除清单」排除 vc"""
    known = set()
    xlsx = config.MAPPING_XLSX
    if os.path.exists(xlsx):
        import openpyxl
        wb = openpyxl.load_workbook(xlsx, data_only=True)
        if "映射总表" in wb.sheetnames:
            ws = wb["映射总表"]
            for r in ws.iter_rows(min_row=2, values_only=True):
                if r[1]:
                    known.add(r[1])
        if "已排除清单" in wb.sheetnames:
            ws = wb["已排除清单"]
            for r in ws.iter_rows(min_row=2, values_only=True):
                if r[0]:
                    known.add(r[0])
        wb.close()
    return known


def calc_new_vcs():
    """其他店铺在架 vc − 已知 − 空商品 = 待审核/自动补录。
    返回 (normal_items, auto_items, shops_meta)"""
    shops_data, shops_meta = products.load_all_shops()
    known = load_known_vcs()
    prefix_map = mapping.load_prefix_map()
    sid_main = mapping.shop_id()

    vc_info = {}
    for sid, rows in shops_data.items():
        if sid == sid_main:
            continue
        for r in rows:
            vc = r.get("vendorCode")
            if not vc or products.is_empty_product(r):
                continue
            p = mapping.price_of(r)
            if vc not in vc_info:
                vc_info[vc] = {"title": r.get("title") or "", "per_shop": {}, "shops": [], "img": r.get("repImg") or ""}
            vc_info[vc]["per_shop"][sid] = {"price": p, "stock": mapping.stock_summary(r)}
            if sid not in vc_info[vc]["shops"]:
                vc_info[vc]["shops"].append(sid)

    normal, auto = [], []
    for vc, info in sorted(vc_info.items()):
        if vc in known:
            continue
        prices = [d["price"] for d in info["per_shop"].values() if d["price"] is not None]
        item = {"vc": vc, "title": info["title"], "price": prices[0] if prices else None,
                "shops": info["shops"], "img": info["img"]}
        m = VC_PREFIX_RE.match(vc or "")
        if m:
            pfx = m.group(1)
            boss = prefix_map.get(pfx)
            if boss:
                item.update({"prefix": pfx, "bossSku": boss["sku"],
                             "bossCn": boss["cn"], "bossDp": boss["dp"]})
                auto.append(item)
                continue
        normal.append(item)
    return normal, auto, shops_meta


def gen_review_html(normal_items, shops_meta):
    boss = mapping.load_boss()
    boss_opts = workbench._boss_opts_html(boss)
    cards = [workbench._review_card_html(i, x, boss_opts) for i, x in enumerate(normal_items, 1)]

    html = '''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8"><title>多店铺待审核（其余4店新商品）</title>
<style>
body{font-family:"Microsoft YaHei",sans-serif;margin:0;background:#f0f2f5;color:#2c3e50}
.bar{position:sticky;top:0;background:#fff;padding:10px 16px;border-bottom:2px solid #8e44ad;z-index:9;display:flex;gap:14px;align-items:center;flex-wrap:wrap}
.bar b{color:#8e44ad}button{padding:6px 16px;cursor:pointer;border-radius:6px;border:1px solid #8e44ad;background:#8e44ad;color:#fff}
button.ghost{background:#fff;color:#8e44ad}
.wrap{padding:16px}
.card{background:#fff;border:1px solid #ddd;border-radius:8px;margin:10px 0;padding:10px 14px;display:flex;justify-content:space-between;align-items:center;gap:12px}
.card.excluded{opacity:.5;background:#f8f8f8}
.card .hd{display:flex;gap:10px;align-items:center;flex:1;min-width:0}
.card .hd .pimg{width:64px;height:64px;object-fit:contain;border-radius:6px;border:1px solid #eee;background:#fafafa}
.card .info{min-width:0}
.card .tt{font-size:13px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;max-width:720px}
.card .vc{color:#7f8c8d;font-size:12px;margin-top:2px}
.card .shop{display:inline-block;background:#eee;border-radius:4px;padding:0 5px;margin-left:4px;font-size:11px}
.card .op select{min-width:300px;padding:6px;border:1px solid #ccc;border-radius:4px}
#out{width:100%;height:140px;font-family:Consolas,monospace;margin-top:8px}
.stats{color:#7f8c8d;font-size:13px}
</style></head>
<body>
<div class="bar">
  <b>多店铺待审核工作台</b>
  <span class="stats">新商品 <b id="total"></b> 个 · 已归属 <b id="mapped"></b> · 已排除 <b id="excl"></b> · 待定 <b id="pend"></b></span>
  <button onclick="exportJson()">导出核对结果 JSON</button>
  <button class="ghost" onclick="resetAll()">重置</button>
</div>
<div class="wrap">
<div class="hint" style="color:#2c3e50">这些是<b>其余店铺在架但映射表中没有</b>的 vendorCode（出现在店已标注）。请为每个选择：归属到某个商品价格表商品（补录）或 非货盘商品（排除）。导出后发我合并进映射表。</div>
CARDS_PLACEHOLDER
<hr><h3>导出结果：</h3><textarea id="out" placeholder="点上方按钮导出"></textarea>
</div>
<script>
const cards = [...document.querySelectorAll('.card')];
document.getElementById('total').textContent = cards.length;
function refresh(){
  let m=0, e=0;
  cards.forEach(cd => {
    const v = cd.querySelector('select').value;
    const isEx = v.startsWith('EXCLUDE');
    cd.classList.toggle('excluded', isEx);
    if (isEx) e++; else if (v) m++;
  });
  document.getElementById('mapped').textContent = m;
  document.getElementById('excl').textContent = e;
  document.getElementById('pend').textContent = cards.length - m - e;
}
function exportJson(){
  const out = [];
  cards.forEach(cd => {
    const v = cd.querySelector('select').value;
    const base = {vc: cd.dataset.vc, title: cd.dataset.title, price: cd.dataset.price, shops: cd.dataset.shops ? cd.dataset.shops.split(',') : []};
    if (!v) { out.push({...base, action: 'pending'}); return; }
    if (v.startsWith('EXCLUDE')) { out.push({...base, action: 'exclude'}); return; }
    const [sku, cn, dp] = v.split('|');
    out.push({...base, action: 'mapped', bossSku: sku, bossCn: cn, bossDp: dp});
  });
  const text = JSON.stringify(out, null, 2);
  document.getElementById('out').value = text;
  downloadJson(text, '多店铺审核.json');
}
function downloadJson(text, filename){
  const blob = new Blob([text], {type: 'application/json;charset=utf-8'});
  const a = document.createElement('a');
  a.href = URL.createObjectURL(blob);
  a.download = filename;
  document.body.appendChild(a); a.click(); document.body.removeChild(a);
  setTimeout(() => URL.revokeObjectURL(a.href), 2000);
  alert('已下载 ' + filename + '（同时显示在上方文本框）');
}
function resetAll(){ cards.forEach(cd => { cd.querySelector('select').value=''; }); refresh(); }
refresh();
</script></body></html>'''
    html = html.replace("CARDS_PLACEHOLDER", "\n".join(cards))
    workbench._write(config.OUT_REVIEW_HTML, html)
    print(f"待审核工作台已生成：{config.OUT_REVIEW_HTML}（{len(normal_items)} 个新商品）")


def build_shop_coverage():
    """从全部店铺 JSON 构建 shop_coverage: {vc: {sid: {price, stock, title, img}}}"""
    shops_data, shops_meta = products.load_all_shops()
    cov = {}
    for sid, rows in shops_data.items():
        for r in rows:
            vc = r.get("vendorCode")
            if not vc:
                continue
            cov.setdefault(vc, {})[sid] = {"price": mapping.price_of(r), "stock": mapping.stock_summary(r),
                                           "title": r.get("title") or "", "img": r.get("repImg") or ""}
    return cov, shops_meta


def merge(review_file=None):
    """增量合并映射表（映射表 xlsx = 唯一状态源）。
    1) 继承旧归属，仅保留 5 店任一在架的 vc（消失即移除）
    2) 追加本次审核 + 前缀自动补录
    3) 继承「已排除清单」+ 本次排除 → 重建映射表
    """
    um = []
    if review_file and os.path.exists(review_file):
        um = json.load(open(review_file, encoding="utf-8"))
    elif review_file:
        print(f"[警告] 审核文件不存在：{review_file}，按无新审核执行增量同步")

    shops_data, shops_meta = products.load_all_shops()
    alive = set()
    for sid, rows in shops_data.items():
        for r in rows:
            vc = r.get("vendorCode")
            if vc:
                alive.add(vc)
    missing_shops = [s["id"] for s in shops_meta if s["id"] not in shops_data]
    skip_prune = bool(missing_shops)
    if skip_prune:
        print(f"[警告] 以下店铺数据缺失：{missing_shops}（可能未拉取或拉取失败）")
        print(f"       为避免误删这些店铺的独有商品，本次跳过「消失即移除」，请先运行 wb.py fetch 后再 merge")
    bcs = mapping.load_bcs()
    boss = mapping.load_boss()
    dp_by_sku = {b["sku"]: b["dp"] for b in boss}

    old_state, old_excluded = mapping.load_mapping_state()
    removed = sorted(vc for vc in old_state if vc not in alive) if not skip_prune else []
    if removed:
        print(f"[消失即移除] {len(removed)} 个商品已不在任何店铺在架，将从映射表移除：")
        for vc in removed:
            print(f"  {vc} | {old_state[vc]['cn']}")
    elif skip_prune:
        print("[消失即移除] 已跳过（店铺数据缺失）")
    excluded = dict(old_excluded)
    if not skip_prune and excluded:
        excl_removed = sorted(vc for vc in excluded if vc not in alive)
        if excl_removed:
            print(f"[排除清单清理] {len(excl_removed)} 个已排除商品已不在任何店铺在架（WB 已删除），从排除清单移除：")
            for vc in excl_removed:
                print(f"  {vc} | {excluded[vc]}")
            excluded = {vc: r for vc, r in excluded.items() if vc in alive}
    elif skip_prune and excluded:
        print(f"[排除清单清理] 已跳过（店铺数据缺失），{len(excluded)} 条排除记录暂保留")

    extra = []
    extra_unmapped = []
    pending = []
    already = set()

    def add_extra(item):
        vcs = [v for v in (item.get("vendorCodes") or []) if v not in already]
        if not vcs:
            return False
        item["vendorCodes"] = vcs
        already.update(vcs)
        extra.append(item)
        return True

    for vc, st in old_state.items():
        if vc in removed:
            continue
        b = next((x for x in boss if x["cn"] == st["cn"]), None)
        dp = st.get("dp")
        if b and b["dp"] is not None:
            dp = b["dp"]
        add_extra({"sku": b["sku"] if b else "", "cn": st["cn"], "dp": dp,
                   "vendorCodes": [vc], "note": "映射表继承"})

    def add_review_item(x, note):
        act = x.get("action")
        if act == "exclude":
            excluded[x["vc"]] = "非货盘商品（人工确认排除）"
        elif act == "mapped":
            sku = x.get("bossSku") or ""
            dp = dp_by_sku.get(sku) or x.get("bossDp")
            add_extra({"sku": sku, "cn": x.get("bossCn") or "", "dp": dp,
                       "vendorCodes": [x["vc"]], "note": note})
        else:
            pending.append(x["vc"])

    for x in um:
        add_review_item(x, "本次审核补录")

    union, _ = mapping.load_shops_union()
    prefix_map = mapping.load_prefix_map()
    _, auto_items, _, _ = mapping.classify_union(union, boss, prefix_map, already)

    auto_records = []
    for x in auto_items:
        if add_extra({"sku": x["bossSku"], "cn": x["bossCn"], "dp": x["bossDp"],
                      "vendorCodes": [x["vc"]], "note": "前缀自动补录"}):
            auto_records.append({"vc": x["vc"], "prefix": x.get("prefix"), "bossSku": x["bossSku"],
                                 "bossCn": x["bossCn"], "price": x["price"], "shops": x["shops"]})
    if auto_records:
        os.makedirs(config.STATE_DIR, exist_ok=True)
        with open(config.AUTO_ADD_JSON, "w", encoding="utf-8") as f:
            json.dump(auto_records, f, ensure_ascii=False, indent=2)
        print(f"[前缀自动补录] {len(auto_records)} 个 → 清单已保存：{config.AUTO_ADD_JSON}")
        for r in auto_records:
            print(f'  {r["vc"]} → {r["bossCn"]}（前缀 {r["prefix"]}，店 {r["shops"]}）')

    extra = extra + extra_unmapped

    shop_coverage, shops_meta = build_shop_coverage()
    n_boss, n_pick, n_unmap, n_multi = mapping.build_xlsx(
        [], extra, bcs, excluded_vcs=excluded,
        shop_coverage=shop_coverage, shops_meta=shops_meta)
    print(f"合并完成：{config.MAPPING_XLSX}")
    print(f"  继承 {len(old_state) - len(removed)} · 移除 {len(removed)} · 本次补录 {len(um)} · 排除 {len(excluded)}"
          f" · 待定 {len(pending)}")
    print(f"  映射表：商品价格表商品 {n_boss} · 已归属 {n_pick} · 未映射在架 {n_unmap} · 多重映射 {n_multi}")


# ---------------- 入口逻辑（供 cli 调用） ----------------
def run_review():
    normal, auto, shops_meta = calc_new_vcs()
    print(f"其他店铺新商品: 普通 {len(normal)} 个 · 前缀自动补录 {len(auto)} 个")
    for x in auto:
        print(f'  [自动] {x["vc"]} → {x["bossCn"]}（前缀 {x["prefix"]}，店 {x["shops"]}）')
    gen_review_html(normal, shops_meta)


def run_merge(review_file=None):
    merge(review_file)
