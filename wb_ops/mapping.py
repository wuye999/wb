# -*- coding: utf-8 -*-
"""
wb_ops 商品价格表 ↔ 店铺商品映射表构建（原 build_mapping.py 的数据逻辑）

数据职责：商品价格表解析、前缀映射、5 店并集分类、8-Sheet 映射表生成。
HTML 工作台渲染已抽到 workbench.py。
"""
import glob
import json
import math
import os
import re
from collections import defaultdict

import openpyxl

from . import bcs
from . import config
from . import keywords
from . import workbench
# ---------------- 主店 / 常量（懒加载，避免 import 即联网） ----------------
_shop_id_cache = None


def shop_id():
    global _shop_id_cache
    if _shop_id_cache is None:
        _shop_id_cache = config.MAIN_SHOP or bcs.get_main_shop()
    return _shop_id_cache


IMG_COL = 9  # 主图链接列（做超链接）
NMID_COL = 16  # WB商品码列（做超链接）


def detail_headers():
    """映射总表 16 列（13 主数据 + 创建时间/更新时间/WB商品码） + 店铺覆盖列"""
    return [
        '产品中文名', 'vendorCode', '双倍售价', f'店铺{shop_id()}价格(CNY)', '折扣%', 'club折扣%',
        '库存', '俄文标题', '主图链接', '尺寸长(cm)', '尺寸宽(cm)', '尺寸高(cm)', '毛重(kg)',
        '创建时间', '更新时间', 'WB商品码',
    ]


# ---------------- 数据读取 ----------------
def _resolve_dp(formula, ws, row_idx):
    """解析 D 列公式（如 =K57*2）：读目标单元格值 × 系数。失败返回 None"""
    m = re.match(r'=([A-Z]+)(\d+)\s*\*\s*(\d+(?:\.\d+)?)', str(formula))
    if m:
        col = openpyxl.utils.column_index_from_string(m.group(1))
        target = ws.cell(row=int(m.group(2)), column=col).value
        if isinstance(target, (int, float)):
            return target * float(m.group(3))
    return None


def load_boss():
    """商品价格表 → [{idx, sku, cn, dp(双倍售价), img, floor, prefix}]
    新表 7 列：图片/SKU/中文名/双倍售价/尺寸/最低售价/前缀码。
    支持 D 列公式（=F{行}*2 等）；NOTEBOOK 为普通商品行（前缀 NBTB），与其他一致。"""
    wb = openpyxl.load_workbook(config.BOSS_XLSX, data_only=False)
    ws = wb["Sheet1"]
    items = []
    for row in ws.iter_rows(min_row=2):
        sku = str(row[1].value).strip() if row[1].value else ""
        cn = str(row[2].value).strip() if row[2].value else ""
        if not sku and not cn:
            continue
        dp = row[3].value
        if isinstance(dp, str) and dp.startswith("="):
            dp = _resolve_dp(dp, ws, row[0].row)
        prefix = str(row[6].value).strip().upper() if len(row) > 6 and row[6].value else ""
        items.append({
            "idx": len(items) + 1,
            "sku": sku,
            "cn": cn,
            "dp": dp,
            "img": str(row[0].value).strip() if row[0].value else "",
            "floor": int(dp) if dp is not None else None,
            "prefix": prefix,
        })
    wb.close()
    return items


def load_prefix_map():
    """商品价格表「vendorCode前缀码」列 → {prefix: {sku, cn, dp}}（含 NOTEBOOK 行，prefix 非空才登记）
    重复前缀 → 控制台告警（保留首个，不覆盖）"""
    wb = openpyxl.load_workbook(config.BOSS_XLSX, data_only=False)
    ws = wb["Sheet1"]
    pmap = {}
    for row in ws.iter_rows(min_row=2):
        sku = str(row[1].value).strip() if row[1].value else ""
        cn = str(row[2].value).strip() if row[2].value else ""
        if not sku and not cn:
            continue
        prefix = str(row[6].value).strip().upper() if len(row) > 6 and row[6].value else ""
        if not prefix:
            continue
        dp = row[3].value
        if isinstance(dp, str) and dp.startswith("="):
            dp = _resolve_dp(dp, ws, row[0].row)
        if prefix in pmap:
            print(f"[告警] 前缀 {prefix} 重复登记：{pmap[prefix]['sku']} 与 {sku}，保留首个")
            continue
        pmap[prefix] = {"sku": sku, "cn": cn, "dp": dp}
    wb.close()
    return pmap


def apply_latest_dp(result_list):
    """用商品价格表最新双倍售价覆盖核对结果的 doublePrice（保证与商品价格表一致）"""
    boss = load_boss()
    dp_by_sku = {b["sku"]: b["dp"] for b in boss}
    n = 0
    for item in result_list:
        sku = item.get("sku")
        if sku in dp_by_sku and dp_by_sku[sku] is not None:
            if item.get("doublePrice") != dp_by_sku[sku]:
                item["doublePrice"] = dp_by_sku[sku]
                n += 1
    return result_list, n


def load_mapping_state():
    """读现有映射表 xlsx → 增量状态（归属关系 + 已排除清单）。
    返回 (state, excluded)：state={vc:{cn,dp,shop_price,discount,nmId}}；excluded={vc:原因}。映射表不存在 → 空。
    shop_price = 「店铺{id}价格(CNY)」列（任一店铺价列，取第一个）；discount = 「折扣%」列；
    nmId = 「WB商品码」列（默认店 nmId，作上架可靠标识；缺列时为 None）。无列时 None。"""
    state, excluded = {}, {}
    if not os.path.exists(config.MAPPING_XLSX):
        print(f"⚠ [警告] 映射表不存在：{config.MAPPING_XLSX}，将按空状态处理（merge 会重建，历史归属/排除将丢失）")
        return state, excluded
    wb = openpyxl.load_workbook(config.MAPPING_XLSX, data_only=True)
    if "映射总表" in wb.sheetnames:
        ws = wb["映射总表"]
        headers = [str(c.value or "") for c in ws[1]]
        if "vendorCode" in headers:
            i_cn, i_vc = headers.index("产品中文名"), headers.index("vendorCode")
            i_dp = headers.index("双倍售价") if "双倍售价" in headers else None
            i_sp = next((i for i, h in enumerate(headers)
                         if re.match(r"^店铺\d+价格\(CNY\)$", h)), None)
            i_disc = headers.index("折扣%") if "折扣%" in headers else None
            i_nm = headers.index("WB商品码") if "WB商品码" in headers else None
            for r in ws.iter_rows(min_row=2, values_only=True):
                vc = r[i_vc]
                if vc:
                    state[vc] = {"cn": r[i_cn] or "", "dp": r[i_dp] if i_dp is not None else None,
                                 "shop_price": r[i_sp] if i_sp is not None else None,
                                 "discount": r[i_disc] if i_disc is not None else None,
                                 "nmId": (r[i_nm] if i_nm is not None else None)}
    if "已排除清单" in wb.sheetnames:
        ws2 = wb["已排除清单"]
        for r in ws2.iter_rows(min_row=2, values_only=True):
            if r[0]:
                excluded[r[0]] = r[1] or "非货盘商品（人工排除）"
    wb.close()
    return state, excluded


def load_bcs():
    """主店 JSON 在架商品，返回完整字段（附加 vc/wbnm/price/img 便捷字段）"""
    d = json.load(open(config.shop_json_path(shop_id()), encoding="utf-8"))
    rows = [r for r in d["rows"] if not r.get("trashedAt")]
    out = []
    for r in rows:
        sl = r.get("sizeList") or []
        price = sl[0].get("price") if sl else None
        if price is None:
            continue
        c = dict(r)
        c["price"] = int(price)
        c["vc"] = r.get("vendorCode") or ""
        c["wbnm"] = str(c["vc"]).rsplit("-", 1)[-1]
        c["img"] = r.get("repImg") or ""
        out.append(c)
    return out


def size_summary(r):
    """sizeList 压成 'chrtId(techSize)@price; ...'"""
    return "; ".join(f"{s.get('chrtId')}({s.get('techSizeName') or s.get('name') or ''})@{s.get('price')}"
                     for s in (r.get("sizeList") or []))


def stock_summary(r):
    """库存数量：'999'；多规格/多仓库不同数量用逗号分隔（去掉 warehouseId 前缀）"""
    seen, parts = set(), []
    for s in (r.get("sizeList") or []):
        for st in (s.get("stockList") or []):
            am = st.get("amount")
            if am is not None and am not in seen:
                seen.add(am)
                parts.append(str(am))
    return ", ".join(parts)


def price_of(r):
    """商品代表价 = sizeList[0].price"""
    sl = r.get("sizeList") or []
    return sl[0].get("price") if sl else None


# ---------------- 候选计算 ----------------
def load_shops_union():
    """5 店在架并集，按 vendorCode 去重合并（自包含读 JSON，避免循环 import）。
    返回 (union, shops_meta)：
      union: {vc: {'vc','title','price','img','per_shop':{sid:{'price','stock'}},'shops':[sid]}}
      过滤 trashedAt / 无价格（含空商品三缺）；同 vc 跨店合并，代表价取主店优先否则最小 sid 店。"""
    union = {}
    shops_meta = []
    try:
        meta = json.load(open(config.STATUS_JSON, encoding="utf-8"))
        shops_meta = meta.get("shops", []) or []
    except Exception:
        shops_meta = []
    if not shops_meta:  # 状态文件缺失 → 扫描目录推断店铺
        for p in sorted(glob.glob(config.shop_json_path("*"))):
            try:
                sid = int(os.path.basename(p).replace("shop", "").replace("_products_all.json", ""))
                shops_meta.append({"id": sid, "name": f"shop{sid}"})
            except ValueError:
                continue
    sid_main = shop_id()
    for s in shops_meta:
        sid = s["id"]
        p = config.shop_json_path(sid)
        if not os.path.exists(p):
            continue
        d = json.load(open(p, encoding="utf-8"))
        for r in d.get("rows", []):
            if r.get("trashedAt"):
                continue
            sl = r.get("sizeList") or []
            if not sl or sl[0].get("price") is None:
                continue  # 无价格（含空商品）不计
            vc = r.get("vendorCode")
            if not vc:
                continue
            price = sl[0].get("price")
            u = union.setdefault(vc, {"vc": vc, "title": r.get("title") or "", "img": r.get("repImg") or "",
                                      "per_shop": {}, "shops": [], "price": None})
            u["per_shop"][sid] = {"price": price, "stock": stock_summary(r),
                                  "nmId": r.get("nmId"), "createAt": r.get("createAt"),
                                  "updateAt": r.get("updateAt")}
            if sid not in u["shops"]:
                u["shops"].append(sid)
            if u["price"] is None or sid == sid_main:
                u["price"] = price
    return union, shops_meta


def classify_union(union, boss, prefix_map, known):
    """5 店并集 → 四分类（互斥，零重复）：
      cand_vcs:    价格命中候选池（代表价 ∈ 某商品 floor/floor+1）→ 上半区勾选
      auto_items:  前缀命中（中段 4 字母在商品价格表前缀码）→ merge 自动补录
      normal_items: 其余未归属 vc → 下半区卡片选归属
      skipped:     known（映射总表+已排除清单）跳过数
    返回 (cand_vcs, auto_items, normal_items, skipped)"""
    boss_floors = {b["floor"] for b in boss if b.get("floor") is not None}
    cand_vcs, auto_items, normal_items, skipped = [], [], [], 0
    for vc, u in sorted(union.items()):
        if vc in known:
            skipped += 1
            continue
        prices = [d["price"] for d in u["per_shop"].values() if d["price"] is not None]
        m = re.match(config.VC_PREFIX_RE, vc or "")
        boss_row = prefix_map.get(m.group(1)) if m else None
        if boss_row:
            auto_items.append({"vc": vc, "title": u["title"], "price": prices[0] if prices else None,
                               "shops": u["shops"], "prefix": m.group(1),
                               "bossSku": boss_row["sku"], "bossCn": boss_row["cn"], "bossDp": boss_row["dp"]})
            continue
        if u.get("price") is not None:
            p_int = int(u["price"])
            if p_int in boss_floors or (p_int - 1) in boss_floors:
                cand_vcs.append(vc)  # 价格命中 {floor, floor+1} → 上半区
                continue
        normal_items.append({"vc": vc, "title": u["title"], "price": prices[0] if prices else None,
                             "shops": u["shops"], "img": u["img"]})
    return cand_vcs, auto_items, normal_items, skipped


def known_vcs_from_mapping():
    """映射表已知 vc = 映射总表归属 + 已排除清单"""
    state, excluded = load_mapping_state()
    return set(state) | set(excluded)


def build_groups(boss, bcs):
    """按 floor 价分组；返回 (conflict_groups, single_groups, todo_list)"""
    by_floor = defaultdict(list)
    for b in boss:
        if b["floor"] is not None:
            by_floor[b["floor"]].append(b)

    price_index = defaultdict(list)
    for c in bcs:
        price_index[c["price"]].append(c)

    def cands_for(floor_p):
        """候选 = 价格 ∈ {floor_p, floor_p+1} 的商品，附关键词命中数"""
        pool = price_index.get(floor_p, []) + price_index.get(floor_p + 1, [])
        seen, out = set(), []
        for c in pool:
            if c["vc"] in seen:
                continue
            seen.add(c["vc"])
            c = dict(c)
            c["hit"] = 0
            out.append(c)
        return out

    def with_hits(b, cands):
        kws = keywords.keywords_for(b["cn"])
        pats = [re.compile(k, re.I) for k in kws]
        for c in cands:
            c["hit"] = sum(1 for p in pats if p.search(c["title"]))
        return cands

    conflict_groups, single_groups, todo = [], [], []
    for floor_p, bosses in sorted(by_floor.items()):
        cands = cands_for(floor_p)
        if not cands:  # 价格带内无任何店铺商品 → 待核查
            for b in bosses:
                todo.append(b)
            continue
        for b in bosses:
            with_hits(b, cands)
        if len(bosses) > 1:
            conflict_groups.append({"floor": floor_p, "bosses": bosses, "cands": cands})
        else:
            single_groups.append({"floor": floor_p, "bosses": bosses, "cands": cands})
    return conflict_groups, single_groups, todo


# ---------------- 自动预匹配 ----------------
def auto_match(result_list, bcs):
    """自动预匹配（v2，冲突检测）。返回 {idx: [vc, ...]}"""
    price_index = defaultdict(list)
    for c in bcs:
        price_index[c["price"]].append(c)

    hits = {}
    for item in result_list:
        dp = item.get("doublePrice")
        if dp is None:
            continue
        floor = int(dp)
        pool = price_index.get(floor, []) + price_index.get(floor + 1, [])
        kws = keywords.keywords_for(item.get("cn") or "")
        pats = [re.compile(k, re.I) for k in kws]
        s = {c["vc"] for c in pool if any(p.search(c["title"]) for p in pats)}
        if s:
            hits[item["idx"]] = s

    vc_owner = defaultdict(list)
    for idx, s in hits.items():
        for vc in s:
            vc_owner[vc].append(idx)
    conflict_vcs = {vc for vc, owners in vc_owner.items() if len(owners) > 1}

    out = {}
    for idx, s in hits.items():
        clean = [vc for vc in s if vc not in conflict_vcs]
        if clean:
            out[idx] = clean
    return out


# ---------------- 映射表行 / 8-Sheet 构建 ----------------
def row_for(item, vc, c, match, status):
    """生成一行映射数据（16 列）。item 的售价键兼容 doublePrice（核对JSON）与 dp（补充条目）"""
    dp = item.get("doublePrice", item.get("dp"))
    if c is None:
        return [item.get("cn"), vc, dp, "", "", "", "", "", "", "", "", "", "", "", "", ""]
    return [
        item.get("cn"), vc, dp, c["price"],
        c.get("discount"), c.get("clubDiscount"), c.get("_stock") or stock_summary(c),
        c.get("title"), c.get("repImg") or "",
        c.get("dimensionsLength"), c.get("dimensionsWidth"), c.get("dimensionsHeight"),
        c.get("dimensionsWeightBrutto"),
        c.get("createAt"), c.get("updateAt"), c.get("nmId"),
    ]


def build_xlsx(result_list, extra_entries, bcs, excluded_vcs=None, shop_coverage=None, shops_meta=None):
    """生成映射表 xlsx（8 Sheet）。"""
    excluded_vcs = excluded_vcs or set()
    shop_coverage = shop_coverage or {}
    shops_meta = shops_meta or [{"id": shop_id(), "name": f"shop{shop_id()}"}]
    bcs_by_vc = {c["vc"]: c for c in bcs}
    sid_main = shop_id()
    for c in bcs:
        shop_coverage.setdefault(c["vc"], {})[sid_main] = {
            "price": c["price"], "stock": stock_summary(c),
            "nmId": c.get("nmId"), "createAt": c.get("createAt"), "updateAt": c.get("updateAt")}
    picked = set()
    for item in result_list:
        picked.update(item.get("vendorCodes") or [])
    for item in extra_entries:
        picked.update(item.get("vendorCodes") or [])

    vc2boss = defaultdict(list)
    for item in list(result_list) + list(extra_entries):
        for vc in item.get("vendorCodes") or []:
            vc2boss[vc].append(item.get("cn") or "")
    multi = {vc: bs for vc, bs in vc2boss.items() if len(bs) > 1}

    vc_cn = {}
    for item in list(result_list) + list(extra_entries):
        for vc in item.get("vendorCodes") or []:
            vc_cn.setdefault(vc, item.get("cn") or "")

    def rep(c, key, cov):
        """主店优先、其他店兜底的代表值（c 自带优先，否则从 cov 按主店→首店回退）"""
        v = c.get(key) if c is not None else None
        if v is None and cov:
            d = cov.get(sid_main) or next(iter(cov.values()))
            v = d.get(key)
        return v

    def enrich(vc, c):
        cov = shop_coverage.get(vc, {})
        if c is None:
            if not cov:
                return None
            d0 = next((d for d in cov.values() if d.get("price") is not None), None)
            if d0 is None:
                d0 = next(iter(cov.values()))
            return {"vc": vc, "title": d0.get("title") or "", "price": d0.get("price"),
                    "repImg": d0.get("img") or "", "img": d0.get("img") or "",
                    "_stock": d0.get("stock") or "",
                    "nmId": rep(None, "nmId", cov), "createAt": rep(None, "createAt", cov),
                    "updateAt": rep(None, "updateAt", cov)}
        c = dict(c)
        for sid, d in cov.items():
            if sid == sid_main:
                continue
            if c.get("price") is None and d.get("price") is not None:
                c["price"] = d["price"]
            if not stock_summary(c) and d.get("stock"):
                c["_stock"] = d["stock"]
        for k in ("nmId", "createAt", "updateAt"):
            if c.get(k) is None:
                c[k] = rep(c, k, cov)
        return c

    headers = detail_headers()
    wb = openpyxl.Workbook()
    # Sheet1 映射总表
    ws1 = wb.active
    ws1.title = "映射总表"
    ws1.append(headers + ["店铺覆盖"])
    for item in result_list:
        vcs = item.get("vendorCodes") or []
        if not vcs:
            ws1.append(row_for(item, "", None, "", "未匹配") + [""])
            continue
        match = "人工确认" if not item.get("auto") else "自动匹配(参考)"
        for vc in vcs:
            c = enrich(vc, bcs_by_vc.get(vc))
            status = "多重映射冲突" if vc in multi else "已匹配"
            cov = ";".join(str(sid) for sid in sorted(shop_coverage.get(vc, {}).keys())) if vc else ""
            ws1.append(row_for(item, vc, c, match, status) + [cov])
            if c:
                ws1.cell(ws1.max_row, IMG_COL).hyperlink = c.get("repImg") or ""
                if c.get("nmId"):
                    ws1.cell(ws1.max_row, NMID_COL).hyperlink = \
                        f"https://www.wildberries.ru/catalog/{c['nmId']}/detail.aspx?targetUrl=GP"
    for item in extra_entries:
        status = "已补录售价" if item.get("dp") is not None else "待补填售价"
        for vc in item.get("vendorCodes") or []:
            c = enrich(vc, bcs_by_vc.get(vc))
            cov = ";".join(str(sid) for sid in sorted(shop_coverage.get(vc, {}).keys())) if vc else ""
            ws1.append(row_for(item, vc, c, "补充条目(价格带识别)", status) + [cov])
            if c:
                ws1.cell(ws1.max_row, IMG_COL).hyperlink = c.get("repImg") or ""
                if c.get("nmId"):
                    ws1.cell(ws1.max_row, NMID_COL).hyperlink = \
                        f"https://www.wildberries.ru/catalog/{c['nmId']}/detail.aspx?targetUrl=GP"
    ws1.freeze_panes = "A2"
    SIMPLE_WIDTHS = [26, 28, 10, 14, 8, 12, 12, 55, 45, 10, 10, 10, 10, 18, 18, 14, 18]
    for i, w in enumerate(SIMPLE_WIDTHS, 1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Sheet2 多重映射冲突
    ws2 = wb.create_sheet("多重映射冲突")
    ws2.append(["vendorCode", "被勾选的商品价格表商品数", "涉及的商品价格表商品", "处理建议"])
    if multi:
        for vc, bs in multi.items():
            ws2.append([vc, len(bs), "；".join(bs),
                        "同一vendorCode被多个商品勾选 → 需人工裁决归属（或同款确实共用一个vendorCode时确认保留）"])
    else:
        ws2.append(["", "", "", "无多重映射冲突（本次核对干净）"])
    for i, w in enumerate([26, 16, 60, 50], 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Sheet3 未映射商品
    ws3 = wb.create_sheet("未映射商品")
    ws3.append(headers + ["未映射原因"])
    missed = [c for c in bcs if c["vc"] not in picked]
    missed.sort(key=lambda c: c["price"])
    for c in missed:
        if c["vc"] in excluded_vcs:
            reason = "已人工标记为非货盘商品（排除，不参与映射）"
        else:
            reason = "店铺在架但未映射 → 漏配商品候选或非货盘商品，需人工核对"
        row = row_for({"sku": "", "cn": "（未映射）", "doublePrice": None}, c["vc"], c, "", reason)
        ws3.append(row + [reason])
        ws3.cell(ws3.max_row, IMG_COL).hyperlink = c.get("repImg") or ""
    for i, w in enumerate(SIMPLE_WIDTHS + [40], 1):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # 已排除清单
    ws_ex = wb.create_sheet("已排除清单")
    ws_ex.append(["vendorCode", "排除原因", "说明"])
    if isinstance(excluded_vcs, dict):
        excl_items = sorted(excluded_vcs.items())
    else:
        excl_items = [(vc, "非货盘商品（人工排除）") for vc in sorted(excluded_vcs)]
    for vc, reason in excl_items:
        ws_ex.append([vc, reason, "不在商品价格表货盘，确认排除；此清单为增量 merge 的排除状态来源"])
    if not excl_items:
        ws_ex.append(["", "", "无已排除商品"])
    for i, w in enumerate([28, 30, 45], 1):
        ws_ex.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Sheet4 待核查清单
    ws4 = wb.create_sheet("待核查清单")
    ws4.append(["商品SKU", "产品中文名", "双倍售价", "问题说明"])
    picked_cn = set()
    for item in list(result_list) + list(extra_entries):
        if item.get("vendorCodes"):
            picked_cn.add(item.get("cn"))
    boss = load_boss()
    for b in boss:
        if b["cn"] not in picked_cn:
            ws4.append([b["sku"], b["cn"], b["dp"],
                        "商品价格表有该商品但映射表无归属 vendorCode → 需人工核对补录（可能未上架/价格特殊/漏配）"])
    for item in extra_entries:
        if item.get("dp") is None:
            note = (item.get("note", "") or "") + " → 需补填双倍售价并确认型号分组"
            ws4.append([item.get("sku"), item.get("cn"), item.get("dp"), note])
    for i, w in enumerate([14, 30, 10, 70], 1):
        ws4.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Sheet5 店铺全量商品
    ws5 = wb.create_sheet("店铺全量商品")
    ws5.append(headers + ["已在映射表"])
    for c in sorted(bcs, key=lambda c: c["price"]):
        row = row_for({"sku": "", "cn": "（店铺商品）", "doublePrice": None}, c["vc"], c, "", "")
        row.append("是" if c["vc"] in picked else "")
        ws5.append(row)
        ws5.cell(ws5.max_row, IMG_COL).hyperlink = c.get("repImg") or ""
    for i, w in enumerate(SIMPLE_WIDTHS + [10], 1):
        ws5.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Sheet6 店铺覆盖矩阵
    ws6 = wb.create_sheet("店铺覆盖矩阵")
    shop_ids = [s["id"] for s in shops_meta]
    ws6.append(["vendorCode", "产品中文名"] + [f"{s['id']}({s['name']})" for s in shops_meta])
    for vc in sorted(picked):
        row = [vc, vc_cn.get(vc, "")]
        for sid in shop_ids:
            d = shop_coverage.get(vc, {}).get(sid)
            row.append(d.get("stock", "") if d else "")
        ws6.append(row)
    for i, w in enumerate([28, 26] + [14] * len(shop_ids), 1):
        ws6.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws6.freeze_panes = "C2"

    # Sheet7 多店价格一致性
    ws7 = wb.create_sheet("多店价格一致性")
    ws7.append(["vendorCode", "产品中文名", f"主店{sid_main}价格", "店铺ID", "该店价格", "说明"])
    n_warn = 0
    for vc in sorted(picked):
        main_price = shop_coverage.get(vc, {}).get(sid_main, {}).get("price")
        if main_price is None:
            continue
        for sid, d in shop_coverage.get(vc, {}).items():
            if sid == sid_main:
                continue
            p = d.get("price")
            if p is not None and abs(float(main_price) - float(p)) > 0.01:
                ws7.append([vc, vc_cn.get(vc, ""), main_price, sid, p,
                            "该店价格与主店不一致 → 人工复核（同品跨店价格应相同）"])
                n_warn += 1
    if n_warn == 0:
        ws7.append(["", "", "", "", "", "无价格不一致（全部店铺同品同价 ✓）"])
    for i, w in enumerate([28, 26, 14, 10, 14, 50], 1):
        ws7.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    wb.save(config.MAPPING_XLSX)
    n_unmapped = sum(1 for c in bcs if c["vc"] not in picked)
    return len(result_list), len(picked), n_unmapped, len(multi)


# ---------------- 入口逻辑（供 cli 调用） ----------------
def run_mapping(legacy=False):
    """生成核对工作台。legacy=True 仅主店候选；否则 5 店并集一页两区。"""
    boss = load_boss()

    if legacy:
        bcs_data = load_bcs()
        conflict_groups, single_groups, todo = build_groups(boss, bcs_data)
        n_conf, n_single, n_todo = workbench.render_html(conflict_groups, single_groups, todo)
        print(f"商品价格表商品 {len(boss)} 个 → 冲突组 {n_conf} 个({sum(len(g['bosses']) for g in conflict_groups)} 商品)"
              f" · 普通 {n_single} 个 · 待核查 {n_todo} 个: {[b['cn'] for b in todo]}")
        print(f"HTML 工作台已生成：{config.OUT_MAPPING_HTML}（legacy：仅主店）")
        return

    union, shops_meta = load_shops_union()
    if not union:
        raise RuntimeError("未找到任何店铺数据，请先运行 wb.py fetch")
    known = known_vcs_from_mapping()
    prefix_map = load_prefix_map()
    cand_vcs, auto_items, normal_items, skipped = classify_union(union, boss, prefix_map, known)
    cand_items = [union[vc] for vc in cand_vcs]
    conflict_groups, single_groups, todo = build_groups(boss, cand_items)
    stats = {"cand": len(cand_vcs), "normal": len(normal_items),
             "auto": len(auto_items), "known": skipped,
             "shops": [s["id"] for s in shops_meta]}
    n_conf, n_single, n_todo, n_normal = workbench.render_unified_html(
        conflict_groups, single_groups, todo, normal_items, shops_meta, stats)

    n_union = len(union)
    n_sum = len(cand_vcs) + len(normal_items) + len(auto_items)
    ident = "✓" if n_union - skipped == n_sum else "✗"
    print(f"5 店并集 {n_union} · 已知(映射表)跳过 {skipped} · 候选池 {len(cand_vcs)} · 未归属新 vc {len(normal_items)}"
          f" · 前缀自动 {len(auto_items)}")
    print(f"  恒等式 |union|-|known|==cand+normal+auto：{n_union}-{skipped}={n_sum} {ident}")
    print(f"  商品价格表商品 {len(boss)} → 冲突组 {n_conf} · 普通 {n_single} · 待核查 {n_todo} · 下半区卡片 {n_normal}")
    print(f"HTML 统一工作台已生成：{config.OUT_MAPPING_HTML}（导出 vc 中心格式 → wb.py merge 统一审核.json）")


def import_mapping(import_file, auto=False):
    """导入核对结果 JSON → 生成映射表 xlsx（旧格式，单店初建）。"""
    result = json.load(open(import_file, encoding="utf-8"))
    if isinstance(result, dict) and "rows" in result:
        result = result["rows"]
    bcs_data = load_bcs()
    result, n_dp = apply_latest_dp(result)
    if n_dp:
        print(f"[双倍售价同步] 覆盖 {n_dp} 个商品为商品价格表最新值")
    if auto:
        auto_map = auto_match(result, bcs_data)
        n_auto = 0
        for item in result:
            if not item.get("vendorCodes") and item["idx"] in auto_map:
                item["vendorCodes"] = auto_map[item["idx"]]
                item["auto"] = True
                n_auto += 1
        print(f"[自动预匹配] 补充 {n_auto} 个商品（仅参考，非正式映射）")
    n_boss, n_pick, n_unmap, n_multi = build_xlsx(result, [], bcs_data)
    print(f"映射表已生成：{config.MAPPING_XLSX}")
    print(f"  商品价格表商品 {n_boss} 个 · 勾选 vendorCode {n_pick} 个"
          f" · 未映射在架 {n_unmap} 个 · 多重映射冲突 {n_multi} 个")
