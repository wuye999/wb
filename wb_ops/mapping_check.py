# -*- coding: utf-8 -*-
"""
wb_ops 映射表核查工作台（带图，人工核对匹配是否有误）（原 mapping_check.py）

读映射总表 Sheet，按商品价格表商品分组带图展示，自动标记三类可疑项：
  ① 前缀不符  ② 价格偏差  ③ 跨店不一致
顶部可「只看可疑」「搜索过滤」「导出可疑清单 JSON」。
"""
import glob
import json
import os
import re
from collections import defaultdict
from html import escape

import openpyxl

from . import config
from . import mapping
PRICE_DEV_TOL = 5  # 价格偏差阈值（元）


def load_mapping_rows():
    """读映射总表 Sheet 全部行 → [{cn, vc, dp, price, discount, club, stock, title, img, ...}]"""
    wb = openpyxl.load_workbook(config.MAPPING_XLSX, data_only=True)
    rows = []
    if "映射总表" in wb.sheetnames:
        ws = wb["映射总表"]
        headers = [str(c.value or "") for c in ws[1]]
        idx = {h: i for i, h in enumerate(headers)}
        need = ["产品中文名", "vendorCode"]
        if not all(k in idx for k in need):
            wb.close()
            raise RuntimeError("映射总表缺少必需列（产品中文名/vendorCode），请先 merge 重建映射表")
        price_col = f"店铺{mapping.shop_id()}价格(CNY)"
        if price_col not in idx:
            print(f"⚠ [警告] 映射总表缺少价格列「{price_col}」（主店可能已变化），价格偏差检测将失效，请重新 fetch + merge")
        for r in ws.iter_rows(min_row=2, values_only=True):
            vc = r[idx["vendorCode"]]
            if not vc:
                continue
            rows.append({
                "cn": r[idx["产品中文名"]] or "",
                "vc": str(vc),
                "dp": r[idx["双倍售价"]] if "双倍售价" in idx else None,
                "price": r[idx[price_col]] if price_col in idx else None,
                "discount": r[idx["折扣%"]] if "折扣%" in idx else None,
                "club": r[idx["club折扣%"]] if "club折扣%" in idx else None,
                "stock": r[idx["库存"]] if "库存" in idx else None,
                "title": r[idx["俄文标题"]] if "俄文标题" in idx else "",
                "img": r[idx["主图链接"]] if "主图链接" in idx else "",
                "d_l": r[idx["尺寸长(cm)"]] if "尺寸长(cm)" in idx else None,
                "d_w": r[idx["尺寸宽(cm)"]] if "尺寸宽(cm)" in idx else None,
                "d_h": r[idx["尺寸高(cm)"]] if "尺寸高(cm)" in idx else None,
                "weight": r[idx["毛重(kg)"]] if "毛重(kg)" in idx else None,
                "shops": r[idx["店铺覆盖"]] if "店铺覆盖" in idx else "",
            })
    wb.close()
    return rows


def load_union_prices():
    """5 店并集价格 → {vc: {sid: price}}（用于跨店一致性检测）"""
    out = {}
    for p in sorted(glob.glob(config.shop_json_path("*"))):
        try:
            sid = int(os.path.basename(p).replace("shop", "").replace("_products_all.json", ""))
        except ValueError:
            continue
        d = json.load(open(p, encoding="utf-8"))
        for r in d.get("rows", []):
            if r.get("trashedAt"):
                continue
            sl = r.get("sizeList") or []
            if not sl or sl[0].get("price") is None:
                continue
            vc = r.get("vendorCode")
            if not vc:
                continue
            out.setdefault(vc, {})[sid] = sl[0].get("price")
    return out


def _vc_prefix(vc):
    m = re.match(config.VC_PREFIX_RE, vc or "")
    return m.group(1) if m else None


def analyze(rows, union_prices, boss_prefix=None):
    """为每行计算可疑标记 → (rows, issues)"""
    issues = []
    prefix_map = mapping.load_prefix_map()
    for r in rows:
        flags = []
        pre = _vc_prefix(r["vc"])
        if pre and pre in prefix_map:
            owner = prefix_map[pre]["cn"]
            if owner != r["cn"]:
                r["flag_prefix"] = f"前缀 {pre} 属于「{owner}」，但归属到「{r['cn']}」"
                issues.append({"vc": r["vc"], "cn": r["cn"], "type": "prefix_conflict", "desc": r["flag_prefix"]})
        elif pre:
            r["note_prefix"] = f"前缀 {pre} 未登记（历史随机前缀）"
        if r["dp"] is not None and r["price"] is not None:
            try:
                floor_p = int(float(r["dp"]))
                cur = int(float(r["price"]))
                if abs(cur - floor_p) >= PRICE_DEV_TOL:
                    r["flag_price"] = f"主店价 {cur} ≠ floor双倍售价 {floor_p}"
                    issues.append({"vc": r["vc"], "cn": r["cn"], "type": "price_dev",
                                   "desc": f"{r['flag_price']}（偏差 {cur - floor_p}）"})
            except (TypeError, ValueError):
                pass
        ps = union_prices.get(r["vc"])
        if ps and len(set(ps.values())) > 1:
            r["flag_cross"] = "跨店价格不一致: " + "; ".join(f"{sid}={v}" for sid, v in sorted(ps.items()))
            issues.append({"vc": r["vc"], "cn": r["cn"], "type": "cross_shop", "desc": r["flag_cross"]})
        r["_any_flag"] = any(k.startswith("flag_") for k in r)
    return rows, issues


def render_html(rows, issues):
    groups = defaultdict(list)
    for r in rows:
        groups[r["cn"]].append(r)
    n_suspect = sum(1 for r in rows if r.get("_any_flag"))
    by_type = defaultdict(int)
    for it in issues:
        by_type[it["type"]] += 1

    group_html = []
    for cn, items in sorted(groups.items(), key=lambda kv: (not any(x.get("_any_flag") for x in kv[1]), kv[0])):
        row_html = []
        for r in items:
            img = escape(r["img"] or "")
            flags = []
            if r.get("flag_prefix"):
                flags.append(f'<span class="f f1">⚠ {escape(r["flag_prefix"])}</span>')
            if r.get("flag_price"):
                flags.append(f'<span class="f f2">⚠ {escape(r["flag_price"])}</span>')
            if r.get("flag_cross"):
                flags.append(f'<span class="f f3">⚠ {escape(r["flag_cross"])}</span>')
            if r.get("note_prefix"):
                flags.append(f'<span class="note">{escape(r["note_prefix"])}</span>')
            flag_html = "".join(flags)
            row_html.append(f'''
<tr class="mrow" data-flag="{1 if r.get("_any_flag") else 0}">
  <td class="picktd"><input type="checkbox" class="badpick" data-vc="{escape(r["vc"])}"
       title="勾选 = 货不对板" {"" if not r.get("_pre_checked") else "checked"}></td>
  <td class="imgtd"><img src="{img}" loading="lazy" class="thumb" onclick="showBig('{escape(img)}')"
       onerror="this.style.opacity=.12" title="点击查看大图"></td>
  <td>
    <div class="vc">{escape(r["vc"])}</div>
    <div class="ru">{escape(r["title"] or "")}</div>
    <div class="meta">
      双倍售价 <b class="dp">¥{r["dp"]}</b> · 主店价 <b>¥{r["price"]}</b>
      {f'· 折扣 {r["discount"]}%' if r["discount"] not in (None, "") else ""}
      {f'· club {r["club"]}%' if r["club"] not in (None, "") else ""}
      {f'· 库存 {r["stock"]}' if r["stock"] not in (None, "") else ""}
      {f'· 尺寸 {r["d_l"]}×{r["d_w"]}×{r["d_h"]}cm' if r["d_l"] not in (None, "") else ""}
      {f'· 毛重 {r["weight"]}kg' if r["weight"] not in (None, "") else ""}
    </div>
    <div class="shops">{escape(str(r["shops"] or ""))}</div>
    {flag_html}
  </td>
</tr>''')
        n_item = len(items)
        n_flag = sum(1 for x in items if x.get("_any_flag"))
        badge = f' <span class="grpbadge">⚠ {n_flag} 可疑</span>' if n_flag else ""
        group_html.append(f'''
<div class="grp" data-anyflag="{1 if n_flag else 0}">
  <div class="ghead"><b>{escape(cn)}</b> <span class="cnt">{n_item} 个 vc</span>{badge}</div>
  <table class="gtable">{"".join(row_html)}</table>
</div>''')

    html = '''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8"><title>映射表核查工作台</title>
<style>
body{font-family:"Microsoft YaHei",sans-serif;margin:0;background:#f0f2f5;color:#2c3e50}
.bar{position:sticky;top:0;background:#fff;padding:10px 16px;border-bottom:2px solid #e67e22;z-index:9;display:flex;gap:14px;align-items:center;flex-wrap:wrap}
.bar b{color:#e67e22}button{padding:6px 14px;cursor:pointer;border-radius:6px;border:1px solid #e67e22;background:#e67e22;color:#fff}
button.ghost{background:#fff;color:#e67e22}
input[type=search]{padding:6px 10px;border:1px solid #ccc;border-radius:6px;min-width:260px}
.wrap{padding:16px}
.stats{color:#7f8c8d;font-size:13px}
.grp{background:#fff;border:1px solid #ddd;border-radius:8px;margin:12px 0;padding:10px 14px}
.ghead{display:flex;gap:12px;align-items:baseline;flex-wrap:wrap;margin-bottom:8px}
.ghead b{font-size:15px}.cnt{color:#7f8c8d;font-size:12px}.grpbadge{color:#c0392b;font-size:12px;font-weight:bold}
.gtable{border-collapse:collapse;width:100%}
.gtable td{border:1px solid #eee;vertical-align:top;padding:8px}
.gtable .picktd{width:40px;text-align:center;background:#fafafa}
.gtable .picktd input{width:18px;height:18px;cursor:pointer;accent-color:#c0392b}
.mrow.picked{background:#fdedec}
.mrow.picked .vc{color:#c0392b;font-weight:bold}
.gtable .imgtd{width:110px;text-align:center}
.gtable img{width:96px;height:96px;object-fit:contain;border-radius:6px;background:#fafafa;border:1px solid #eee}
.gtable .thumb{cursor:zoom-in;transition:transform .15s}
.gtable .thumb:hover{transform:scale(1.05);border-color:#e67e22}
#lightbox{position:fixed;inset:0;background:rgba(0,0,0,.85);z-index:9999;display:none;
  align-items:center;justify-content:center;flex-direction:column;cursor:zoom-out}
#lightbox.show{display:flex}
#lb-img{max-width:92vw;max-height:85vh;object-fit:contain;border-radius:8px;background:#fff;box-shadow:0 8px 40px rgba(0,0,0,.6)}
#lb-close{position:absolute;top:14px;right:20px;font-size:34px;color:#fff;cursor:pointer;line-height:1;opacity:.9}
#lb-close:hover{opacity:1;transform:scale(1.15)}
#lb-title{color:#fff;font-size:13px;margin-top:10px;max-width:90vw;text-align:center;word-break:break-all}
.vc{font-family:Consolas,monospace;font-size:12px;color:#7f8c8d}
.ru{font-size:13px;margin:2px 0;color:#2c3e50}
.meta{font-size:12px;color:#555;margin:2px 0}.meta .dp{color:#e74c3c}
.shops{font-size:11px;color:#95a5a6}
.f{display:inline-block;margin:3px 6px 0 0;padding:2px 8px;border-radius:4px;font-size:11px;font-weight:bold}
.f1{background:#fdedec;color:#c0392b}.f2{background:#fef9e7;color:#b7950b}.f3{background:#ebf5fb;color:#2471a3}
.note{display:inline-block;margin:3px 6px 0 0;padding:2px 8px;border-radius:4px;font-size:11px;color:#95a5a6;background:#f5f5f5}
.hint{color:#8e44ad;font-size:12px;margin:6px 0}
</style></head>
<body>
<div class="bar">
  <b>映射表核查工作台</b>
  <span class="stats">共 <b id="total"></b> 条 · <b id="groups"></b> 个商品价格表商品 · 可疑 <b id="suspect" style="color:#c0392b"></b> 条
    · 已勾选货不对板 <b id="picked" style="color:#c0392b"></b> 条</span>
  <input type="search" id="q" placeholder="搜索 vc / 中文名 / 俄文标题…" oninput="applyFilter()">
  <button id="btnSuspect" class="ghost" onclick="toggleSuspect()">只看可疑</button>
  <button id="btnPicked" class="ghost" onclick="togglePicked()">只看已勾选</button>
  <button class="ghost" onclick="exportPicked()">导出勾选 vc (JSON)</button>
  <button class="ghost" onclick="copyVcs()">复制 vc (逗号分隔)</button>
  <button class="ghost" onclick="clearPicked()">清空勾选</button>
</div>
<div class="wrap">
<div class="hint">
  <b>快速筛查货不对板</b>：逐个商品看图片与俄文标题是否与中文名一致，<b>勾选左侧复选框 = 标记货不对板</b>（勾选状态自动保存在浏览器，刷新不丢）。核对完点「导出勾选 vc (JSON)」下载文件，或用「复制 vc」复制逗号分隔列表（可直接用于 <code>wb.py trash --vc ...</code>）。自动标记：
  <span class="f f1">前缀命中他品（可能配错）</span>
  <span class="f f2">价格偏差 ≥ [[TOL]] 元</span>
  <span class="f f3">跨店价格不一致</span>
  <span class="note">历史随机前缀（未登记，仅供参考）</span>
  （点击图片可看大图；可疑组排在前面）
</div>
GROUPS_PLACEHOLDER
</div>
<div id="lightbox" onclick="closeBig(event)">
  <span id="lb-close" onclick="closeBig()">×</span>
  <img id="lb-img" src="" alt="">
  <div id="lb-title"></div>
</div>
<script>
const ISSUES = ISSUES_JSON;
const ROWS = ROWS_JSON;
document.getElementById('total').textContent = TOTAL;
document.getElementById('groups').textContent = GROUPS_N;
document.getElementById('suspect').textContent = SUSPECT_N;
const LS_KEY = 'mapping_check_picked_v2';
let onlySuspect = false, onlyPicked = false;
function loadPicked(){
  try { return new Set(JSON.parse(localStorage.getItem(LS_KEY) || '[]')); } catch(e){ return new Set(); }
}
function savePicked(){
  document.querySelectorAll('.badpick').forEach(c => {
    c.closest('.mrow').classList.toggle('picked', c.checked);
  });
  const set = new Set();
  document.querySelectorAll('.badpick:checked').forEach(c => set.add(c.dataset.vc));
  localStorage.setItem(LS_KEY, JSON.stringify([...set]));
  refreshCount();
}
function refreshCount(){
  const n = document.querySelectorAll('.badpick:checked').length;
  document.getElementById('picked').textContent = n;
}
document.querySelectorAll('.badpick').forEach(c => c.addEventListener('change', savePicked));
const prev = loadPicked();
if (prev.size){
  document.querySelectorAll('.badpick').forEach(c => {
    if (prev.has(c.dataset.vc)){ c.checked = true; c.closest('.mrow').classList.add('picked'); }
  });
  refreshCount();
}
function applyFilter(){
  const q = document.getElementById('q').value.trim().toLowerCase();
  document.querySelectorAll('.grp').forEach(g => {
    let visible = false;
    g.querySelectorAll('.mrow').forEach(tr => {
      const txt = (tr.textContent || '').toLowerCase();
      const okQ = !q || txt.includes(q);
      const okS = !onlySuspect || tr.dataset.flag === '1';
      const okP = !onlyPicked || tr.querySelector('.badpick').checked;
      tr.style.display = (okQ && okS && okP) ? '' : 'none';
      if (okQ && okS && okP) visible = true;
    });
    g.style.display = visible ? '' : 'none';
  });
}
function toggleSuspect(){
  onlySuspect = !onlySuspect;
  document.getElementById('btnSuspect').classList.toggle('ghost', onlySuspect);
  applyFilter();
}
function togglePicked(){
  onlyPicked = !onlyPicked;
  document.getElementById('btnPicked').classList.toggle('ghost', onlyPicked);
  applyFilter();
}
function pickedRows(){
  const out = [];
  document.querySelectorAll('.badpick:checked').forEach(c => {
    const tr = c.closest('.mrow');
    const r = ROWS.find(x => x.vc === c.dataset.vc) || {};
    out.push({vc: c.dataset.vc, cn: r.cn || '', title: r.title || '', shops: r.shops || ''});
  });
  return out;
}
function exportPicked(){
  const rows = pickedRows();
  if (!rows.length){ alert('未勾选任何商品'); return; }
  const blob = new Blob([JSON.stringify(rows, null, 2)], {type:'application/json;charset=utf-8'});
  const a = document.createElement('a');
  a.href = URL.createObjectURL(blob); a.download = '货不对板vc.json'; a.click();
  setTimeout(() => URL.revokeObjectURL(a.href), 2000);
  alert('已下载 货不对板vc.json（' + rows.length + ' 条）');
}
function fallbackCopy(text, done, fail){
  // 兜底：临时可见 textarea（display:none 元素无法形成选区 → 复制为空）
  const ta = document.createElement('textarea');
  ta.value = text;
  ta.style.position = 'fixed';
  ta.style.top = '-9999px';
  ta.style.left = '0';
  document.body.appendChild(ta);
  ta.focus(); ta.select();
  let ok = false;
  try { ok = document.execCommand('copy'); } catch(e){ ok = false; }
  document.body.removeChild(ta);
  ok ? done() : fail();
}
function copyVcs(){
  const vcs = pickedRows().map(r => r.vc);
  if (!vcs.length){ alert('未勾选任何商品'); return; }
  const text = vcs.join(',');
  const done = () => alert('已复制 ' + vcs.length + ' 个 vc（逗号分隔，可直接用于 wb.py trash --vc）');
  const fail = () => alert('复制失败：请用「导出勾选 vc (JSON)」下载后再手动复制');
  if (navigator.clipboard && window.isSecureContext){
    navigator.clipboard.writeText(text).then(done, () => fallbackCopy(text, done, fail));
    return;
  }
  fallbackCopy(text, done, fail);
}
function clearPicked(){
  document.querySelectorAll('.badpick:checked').forEach(c => c.checked = false);
  savePicked(); applyFilter();
}
function showBig(thumbUrl){
  const box = document.getElementById('lightbox');
  const img = document.getElementById('lb-img');
  const ttl = document.getElementById('lb-title');
  const big = thumbUrl.replace('/tm/', '/big/');
  img.onerror = function(){ if (this.src !== thumbUrl) this.src = thumbUrl; };
  img.src = big;
  ttl.textContent = big;
  box.classList.add('show');
}
function closeBig(ev){
  if (ev && ev.target.id !== 'lightbox' && ev.target.id !== 'lb-close') return;
  document.getElementById('lightbox').classList.remove('show');
  document.getElementById('lb-img').src = '';
}
document.addEventListener('keydown', e => { if (e.key === 'Escape') closeBig(); });
function exportIssues(){
  const blob = new Blob([JSON.stringify(ISSUES, null, 2)], {type:'application/json;charset=utf-8'});
  const a = document.createElement('a');
  a.href = URL.createObjectURL(blob); a.download = '映射表可疑项.json'; a.click();
}
applyFilter();
</script></body></html>'''
    html = html.replace("[[TOL]]", str(PRICE_DEV_TOL))
    html = html.replace("GROUPS_PLACEHOLDER", "\n".join(group_html) if group_html else '<div class="hint">映射表为空</div>')
    html = html.replace("ISSUES_JSON", json.dumps(issues, ensure_ascii=False))
    html = html.replace("ROWS_JSON", json.dumps(
        [{"vc": r["vc"], "cn": r["cn"], "title": r["title"] or "", "shops": str(r.get("shops") or "")} for r in rows],
        ensure_ascii=False))
    html = html.replace("TOTAL", str(len(rows))).replace("GROUPS_N", str(len(groups))).replace("SUSPECT_N", str(n_suspect))
    workbench_write(config.OUT_MAPPING_CHECK_HTML, html)
    return len(rows), len(groups), n_suspect, dict(by_type)


def workbench_write(path, html):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        f.write(html)


def run(tol=PRICE_DEV_TOL):
    global PRICE_DEV_TOL
    PRICE_DEV_TOL = tol
    print(f"读取映射表 {config.MAPPING_XLSX} ...")
    rows = load_mapping_rows()
    if not rows:
        raise RuntimeError("映射表为空，请先 wb.py merge")
    print(f"  映射条目: {len(rows)}")
    print("读取 5 店价格（跨店一致性检测）...")
    union_prices = load_union_prices()
    print("  分析可疑项 ...")
    rows, issues = analyze(rows, union_prices, None)
    total, groups, suspect, by_type = render_html(rows, issues)
    print(f"\n已生成: {config.OUT_MAPPING_CHECK_HTML}")
    print(f"  共 {total} 条 · {groups} 个商品价格表商品 · 可疑 {suspect} 条")
    for t, n in by_type.items():
        print(f"    {t}: {n}")
    print("  打开 HTML 逐个核对图片/俄文标题与中文名是否一致；可疑项已自动标记并排前。")
